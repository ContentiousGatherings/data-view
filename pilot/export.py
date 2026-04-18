"""
Excel export for pilot dataset.

Builds an .xlsx workbook with a data sheet and embedded charts.
"""

from collections import Counter

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pilot.extract import PilotRow

# Column definitions: (header, dataclass field, column width)
COLUMNS = [
    ("AuthEvent ID", "auth_event_id", 14),
    ("Datum", "datum", 12),
    ("Kategori", "kategori", 16),
    ("Ort", "ort", 20),
    ("Ort ID", "ort_id", 10),
    ("Län", "lan", 18),
    ("Aktörer", "aktorer", 30),
    ("Aktör-IDn", "aktor_idn", 16),
    ("Textutdrag", "textutdrag", 50),
    ("Länkar", "lankar", 40),
    ("Tidning", "tidning", 20),
    ("Käll-event-IDn", "kall_event_idn", 18),
    ("Artikel-IDn", "artikel_idn", 18),
    ("MeetingType-IDn", "meetingtype_idn", 18),
    ("Sökord", "sokord", 16),
    ("Källor", "kallor", 10),
]


def export_xlsx(rows: list[PilotRow], output_path: str) -> None:
    """Export pilot dataset rows to an .xlsx file with charts."""
    wb = Workbook()

    _write_data_sheet(wb, rows)
    _write_year_chart(wb, rows)
    _write_county_chart(wb, rows)

    wb.save(output_path)
    print(f"Wrote {len(rows)} rows to {output_path}")


def _write_data_sheet(wb: Workbook, rows: list[PilotRow]) -> None:
    """Write the main data sheet."""
    ws = wb.active
    ws.title = "Data"

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Write headers
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_, field_name, _) in enumerate(COLUMNS, start=1):
            value = getattr(row, field_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Wrap text for long content
            if field_name in ("textutdrag", "lankar", "aktorer"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"


def _write_year_chart(wb: Workbook, rows: list[PilotRow]) -> None:
    """Write a sheet with events-per-year line chart."""
    ws = wb.create_sheet("Möten per år")

    # Collect all search terms
    all_terms = sorted({t for r in rows for t in (r.sokord or "").split("; ") if t})
    if not all_terms:
        return

    # Count events per year per term
    year_counts: dict[str, Counter] = {term: Counter() for term in all_terms}
    for row in rows:
        if not row.datum:
            continue
        year = row.datum[:4]
        for term in (row.sokord or "").split("; "):
            term = term.strip()
            if term in year_counts:
                year_counts[term][year] += 1

    # Get sorted year range
    all_years = sorted(
        {y for counts in year_counts.values() for y in counts}
    )
    if not all_years:
        return

    # Write summary table
    ws.cell(row=1, column=1, value="År").font = Font(bold=True)
    for i, term in enumerate(all_terms):
        ws.cell(row=1, column=i + 2, value=term).font = Font(bold=True)

    for row_idx, year in enumerate(all_years, start=2):
        ws.cell(row=row_idx, column=1, value=int(year))
        for col_idx, term in enumerate(all_terms, start=2):
            ws.cell(row=row_idx, column=col_idx, value=year_counts[term].get(year, 0))

    # Create chart
    chart = LineChart()
    chart.title = "Möten per år"
    chart.x_axis.title = "År"
    chart.y_axis.title = "Antal"
    chart.style = 10
    chart.width = 25
    chart.height = 15

    # Categories (years)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(all_years) + 1)
    chart.set_categories(cats)

    # Data series
    for i, term in enumerate(all_terms):
        data = Reference(
            ws,
            min_col=i + 2,
            min_row=1,
            max_row=len(all_years) + 1,
        )
        chart.add_data(data, titles_from_data=True)

    # Place chart below the data table
    chart_row = len(all_years) + 4
    ws.add_chart(chart, f"A{chart_row}")

    # Column widths
    ws.column_dimensions["A"].width = 8
    for i in range(len(all_terms)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16


def _write_county_chart(wb: Workbook, rows: list[PilotRow]) -> None:
    """Write a sheet with events-per-county bar chart."""
    ws = wb.create_sheet("Möten per län")

    # Collect all search terms
    all_terms = sorted({t for r in rows for t in (r.sokord or "").split("; ") if t})
    if not all_terms:
        return

    # Count events per county per term
    county_counts: dict[str, Counter] = {term: Counter() for term in all_terms}
    for row in rows:
        county = row.lan or "(okänt)"
        for term in (row.sokord or "").split("; "):
            term = term.strip()
            if term in county_counts:
                county_counts[term][county] += 1

    # Sort counties by total count descending
    all_counties_counter: Counter = Counter()
    for counts in county_counts.values():
        all_counties_counter.update(counts)
    all_counties = [c for c, _ in all_counties_counter.most_common()]

    if not all_counties:
        return

    # Write summary table
    ws.cell(row=1, column=1, value="Län").font = Font(bold=True)
    for i, term in enumerate(all_terms):
        ws.cell(row=1, column=i + 2, value=term).font = Font(bold=True)

    for row_idx, county in enumerate(all_counties, start=2):
        ws.cell(row=row_idx, column=1, value=county)
        for col_idx, term in enumerate(all_terms, start=2):
            ws.cell(
                row=row_idx, column=col_idx, value=county_counts[term].get(county, 0)
            )

    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Möten per län"
    chart.x_axis.title = "Län"
    chart.y_axis.title = "Antal"
    chart.style = 10
    chart.width = 30
    chart.height = 15

    # Categories (counties)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(all_counties) + 1)
    chart.set_categories(cats)

    # Data series
    for i, term in enumerate(all_terms):
        data = Reference(
            ws,
            min_col=i + 2,
            min_row=1,
            max_row=len(all_counties) + 1,
        )
        chart.add_data(data, titles_from_data=True)

    # Place chart below the data table
    chart_row = len(all_counties) + 4
    ws.add_chart(chart, f"A{chart_row}")

    # Column widths
    ws.column_dimensions["A"].width = 22
    for i in range(len(all_terms)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16
